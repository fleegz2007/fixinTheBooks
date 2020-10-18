from openpyxl import load_workbook
from datetime import datetime
from datetime import date
import calendar

print('What is your monthly payment amount?')
monthlypayment = input()

print('What is your weekly paymnet amount?')
weeklypayment = input()

print('Any additional payments?')
additionalpayment = input()

today = datetime.today()
strMonth = None
if today.month < 10:
    strMonth = '0' + str(today.month)
else:
    strMonth = str(today.month)

daysInMonth = calendar.monthrange(today.year, today.month)[1]

journalPath = 'Z:/Personal Accounting/' + str(today.year)[0:4] + '/' + 'TestPath' '/Journal ' + '09' + str(today.year)[2:4] + '.xlsx'
print(journalPath)

print('Calculating revenue accruals...')
dailyWeekly = round((weeklypayment / 7), 2)
dailyadd = round((additionalpayment / daysInMonth), 2)
dailyMonthly = round(((monthlypayment) / daysInMonth), 2)
print('Daily accrual of Fiserv payment: ' + str(dailyMonthly))
print('Daily accrual of Vanguard investment: ' + str(dailyadd))
print('Daily accrual of Rebecca and Nate payment: ' + str(dailyWeekly))
today = date.today()
begMonth = date(today.year, today.month, 1)
timeDif = int(str((today - begMonth))[0:2])

journalwb = load_workbook(filename=journalPath)
sheetNames = journalwb.sheetnames[1:32]

print('Pasting accruals into current months workbook...')

#With this function, journal entry will not overwrite entries if a value exists in Cell A4 of the workbook
def masterbooking():
    i = 0
    while i <= timeDif:
        if journalwb[str(sheetNames[i])]['A4'].value is None:
            journalwb[str(sheetNames[i])]['A4'] = 'Accounts Receivable: Monthly'
            journalwb[str(sheetNames[i])]['B5'] = 'Monthly Revenue'
            journalwb[str(sheetNames[i])]['J4'] = dailyMonthly
            journalwb[str(sheetNames[i])]['L5'] = dailyMonthly
        if journalwb[str(sheetNames[i])]['A7'].value is None:
            journalwb[str(sheetNames[i])]['A7'] = 'Accounts Receivable: Additional'
            journalwb[str(sheetNames[i])]['B8'] = 'Additional Revenue'
            journalwb[str(sheetNames[i])]['J7'] = dailyadd
            journalwb[str(sheetNames[i])]['L8'] = dailyadd
        if journalwb[str(sheetNames[i])]['A10'].value is None:
            journalwb[str(sheetNames[i])]['A10'] = 'Accounts Receivable: Weekly'
            journalwb[str(sheetNames[i])]['B11'] = 'Weekly Revenue'
            journalwb[str(sheetNames[i])]['J10'] = dailyWeekly
            journalwb[str(sheetNames[i])]['L11'] = dailyWeekly
            i += 1
        i += 1
    journalwb.save(journalPath)

masterbooking()








